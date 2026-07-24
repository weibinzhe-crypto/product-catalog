import openpyxl
import json

# 读取产品价目表
xlsx_path = r"D:\Users\zheng\Downloads\整理产品数据\产品价目表_修正名称_20260724_124813.xlsx"
wb = openpyxl.load_workbook(xlsx_path, data_only=True)
ws = wb.active

products = []
for row in range(2, ws.max_row + 1):
    brand = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=2).value
    spec = ws.cell(row=row, column=3).value
    retail_price = ws.cell(row=row, column=4).value
    box_spec = ws.cell(row=row, column=5).value
    barcode = ws.cell(row=row, column=6).value
    price = ws.cell(row=row, column=7).value
    product_code = ws.cell(row=row, column=8).value

    if not name:
        continue

    products.append({
        "brand": str(brand) if brand else "",
        "name": str(name) if name else "",
        "spec": str(spec) if spec else "",
        "retailPrice": float(retail_price) if retail_price else 0,
        "boxSpec": str(box_spec) if box_spec else "",
        "barcode": str(barcode) if barcode else "",
        "price": float(price) if price else 0,
        "productCode": str(product_code) if product_code else "",
        "image": ""
    })

wb.close()

# 保存为JSON
output_path = r"D:\Users\zheng\Downloads\product-catalog\data\products.json"
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print(f"已导出 {len(products)} 条产品数据到 {output_path}")
